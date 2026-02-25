"""
Excel helper for saving applied job data to .xlsx files.
Saves one job at a time immediately after successful application.
"""

import os
from datetime import datetime


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Track if separator was already added for this session
_separator_added = False


def save_to_excel(job):
    """Save a single applied job to excels/applied_jobs_{DD_MM_YYYY}.xlsx.

    If the file already exists (re-run same day), appends data with a blue separator row
    (only once per session).

    Args:
        job: dict with keys: position, company, location, experience, salary, skills, url, applied_at
    """
    global _separator_added

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        print("  ⚠️ openpyxl not installed. Run: pip install openpyxl")
        return

    excels_dir = os.path.join(PROJECT_ROOT, "excels", "naukri")
    os.makedirs(excels_dir, exist_ok=True)

    today_str = datetime.now().strftime("%d_%m_%Y")
    filename = f"applied_jobs_{today_str}.xlsx"
    filepath = os.path.join(excels_dir, filename)

    headers = ["S.No", "Position", "Company", "Location",
               "Experience", "Salary", "Skills", "Link", "Applied At"]

    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active

        # Add blue separator row once per session
        if not _separator_added:
            separator_row = ws.max_row + 2
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=separator_row, column=col_idx, value="")
                cell.fill = blue_fill
            _separator_added = True

        # Determine next S.No
        existing_count = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                existing_count = max(existing_count, int(row[0]))

        sno = existing_count + 1
        write_row = ws.max_row + 1
        _write_job_row(ws, write_row, sno, job, link_font)
        wb.save(filepath)
    else:
        # Create new file
        wb = Workbook()
        ws = wb.active
        ws.title = "Applied Jobs"

        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font

        _write_job_row(ws, 2, 1, job, link_font)
        wb.save(filepath)
        _separator_added = True  # No separator needed for first run

    print(f"    📁 Saved to {filename}")


def _write_job_row(ws, row_num, sno, job, link_font):
    """Write a single job row to the worksheet."""
    url = job.get("url", "")
    row_data = [
        sno,
        job.get("position", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("experience", ""),
        job.get("salary", ""),
        job.get("skills", ""),
        url,
        job.get("applied_at", ""),
    ]

    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_idx, value=val)

    # Make the Link column a clickable hyperlink
    if url:
        link_cell = ws.cell(row=row_num, column=8)  # Column 8 = Link (no more Openings)
        link_cell.hyperlink = url
        link_cell.font = link_font
        link_cell.value = url
