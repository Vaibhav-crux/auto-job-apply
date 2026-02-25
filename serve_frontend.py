"""
Simple HTTP server to serve the Applied Jobs Viewer frontend
and provide JSON API for reading Excel data.

Usage: python serve_frontend.py
Opens: http://localhost:8501
"""

import os
import sys
import json
import glob
import http.server
import socketserver
import webbrowser
from urllib.parse import urlparse, parse_qs

# Ensure project root is in path
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

from config.settings import default_tab

EXCELS_DIR = os.path.join(PROJECT_ROOT, "excels")
FRONTEND_DIR = os.path.join(PROJECT_ROOT, "frontend")
PORT = 8501


def read_excel_folder(folder_path):
    """Read all .xlsx files from a folder and return job data grouped by date."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed"}

    jobs_by_date = {}  # { "25 Feb 2026": [ {job}, {job} ] }

    xlsx_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")), reverse=True)

    for filepath in xlsx_files:
        basename = os.path.basename(filepath)
        # Skip temp files
        if basename.startswith("~$"):
            continue

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Read headers from row 1
            headers = []
            for cell in ws[1]:
                headers.append((cell.value or "").strip())

            link_col_idx = headers.index("Link") if "Link" in headers else -1

            for row in ws.iter_rows(min_row=2):
                values = [cell.value for cell in row]

                # Skip empty/separator rows
                if not values or all(v is None or str(v).strip() == "" for v in values):
                    continue

                # Build job dict from headers
                job = {}
                for i, header in enumerate(headers):
                    if i < len(values):
                        val = values[i]
                        job[header] = str(val).strip() if val is not None else ""
                    else:
                        job[header] = ""

                # Skip if no position
                if not job.get("Position", ""):
                    continue

                # Extract date from "Applied At" (format: "2026-02-25 19:01:15")
                applied_at = job.get("Applied At", "")
                date_key = "Unknown Date"
                time_str = ""

                if applied_at:
                    try:
                        from datetime import datetime
                        dt = datetime.strptime(applied_at, "%Y-%m-%d %H:%M:%S")
                        date_key = dt.strftime("%d %b %Y")  # "25 Feb 2026"
                        time_str = dt.strftime("%I:%M %p")   # "07:01 PM"
                    except (ValueError, TypeError):
                        pass

                job["time"] = time_str

                # Extract hyperlink URL if available
                if link_col_idx >= 0 and link_col_idx < len(row):
                    link_cell = row[link_col_idx]
                    if link_cell.hyperlink and link_cell.hyperlink.target:
                        job["Link"] = str(link_cell.hyperlink.target)

                if date_key not in jobs_by_date:
                    jobs_by_date[date_key] = []
                jobs_by_date[date_key].append(job)

            wb.close()
        except Exception as e:
            print(f"  [!] Error reading {basename}: {e}")

    # Sort dates newest first
    sorted_dates = sorted(
        jobs_by_date.keys(),
        key=lambda d: _parse_date_key(d),
        reverse=True
    )

    result = []
    for date_key in sorted_dates:
        result.append({
            "date": date_key,
            "jobs": jobs_by_date[date_key]
        })

    return result


def _parse_date_key(date_str):
    """Parse '25 Feb 2026' to a sortable value."""
    try:
        from datetime import datetime
        return datetime.strptime(date_str, "%d %b %Y")
    except (ValueError, TypeError):
        from datetime import datetime
        return datetime.min


def get_tabs():
    """Get list of tab folders from excels/ directory."""
    if not os.path.exists(EXCELS_DIR):
        return []
    tabs = []
    for name in sorted(os.listdir(EXCELS_DIR)):
        full_path = os.path.join(EXCELS_DIR, name)
        if os.path.isdir(full_path) and not name.startswith("."):
            tabs.append(name)
    return tabs


class RequestHandler(http.server.SimpleHTTPRequestHandler):
    """Custom handler for serving frontend + API."""

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path

            # API: Get tabs
            if path == "/api/tabs":
                tabs = get_tabs()
                self._json_response({"tabs": tabs, "default": default_tab})
                return

            # API: Get jobs for a tab
            if path.startswith("/api/jobs/"):
                tab_name = path.split("/api/jobs/", 1)[1].strip("/")
                folder_path = os.path.join(EXCELS_DIR, tab_name)
                if not os.path.isdir(folder_path):
                    self._json_response({"error": "Tab not found"}, 404)
                    return
                data = read_excel_folder(folder_path)
                if isinstance(data, dict) and "error" in data:
                    self._json_response(data, 500)
                    return
                self._json_response({"tab": tab_name, "groups": data})
                return

            # Serve frontend files
            if path == "/" or path == "/index.html":
                self._serve_file("applied_jobs_view.html", "text/html")
                return
            if path.endswith(".css"):
                filename = os.path.basename(path)
                self._serve_file(filename, "text/css")
                return
            if path.endswith(".js"):
                filename = os.path.basename(path)
                self._serve_file(filename, "application/javascript")
                return

            self.send_error(404)
        except Exception as e:
            print(f"[!] Request error: {e}")
            try:
                self._json_response({"error": str(e)}, 500)
            except Exception:
                pass

    def _json_response(self, data, status=200):
        """Send JSON response."""
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def _serve_file(self, filename, content_type):
        """Serve a file from the frontend directory."""
        filepath = os.path.join(FRONTEND_DIR, filename)
        if not os.path.exists(filepath):
            self.send_error(404, f"File not found: {filename}")
            return
        self.send_response(200)
        self.send_header("Content-Type", f"{content_type}; charset=utf-8")
        self.end_headers()
        with open(filepath, "rb") as f:
            self.wfile.write(f.read())

    def log_message(self, format, *args):
        """Suppress default access logs."""
        pass


class ReusableTCPServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = True
    daemon_threads = True


def main():
    with ReusableTCPServer(("", PORT), RequestHandler) as httpd:
        url = f"http://localhost:{PORT}"
        print(f"[*] Applied Jobs Viewer running at {url}")
        print(f"    Press Ctrl+C to stop\n")
        webbrowser.open(url)
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n[*] Server stopped.")


if __name__ == "__main__":
    main()
